import pandas as pd
import os
from datetime import datetime
from typing import Dict, Any, Optional, List
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ReportExporter:
    def __init__(self, data: Dict[str, pd.DataFrame], transformed_data: Dict[str, pd.DataFrame], 
                 metrics_calculator, validator):
        self.data = data
        self.transformed_data = transformed_data
        self.metrics = metrics_calculator
        self.validator = validator
        self.filters = self._extract_filters()

    def _extract_filters(self) -> Dict[str, Any]:
        filters = {}
        
        if self.transformed_data.get('registrations') is not None:
            reg_df = self.transformed_data['registrations']
            if 'reg_date' in reg_df.columns:
                date_series = pd.to_datetime(reg_df['reg_date'], errors='coerce')
                valid_dates = date_series.dropna()
                if len(valid_dates) > 0:
                    filters['start_date'] = valid_dates.min().strftime('%Y-%m-%d')
                    filters['end_date'] = valid_dates.max().strftime('%Y-%m-%d')
            
            if 'department_name' in reg_df.columns:
                depts = reg_df['department_name'].dropna().unique().tolist()
                filters['departments'] = depts if len(depts) <= 10 else f'共{len(depts)}个科室'
            
            if 'doctor_name' in reg_df.columns:
                doctors = reg_df['doctor_name'].dropna().unique().tolist()
                filters['doctors'] = doctors if len(doctors) <= 10 else f'共{len(doctors)}位医生'
            
            if 'patient_type' in reg_df.columns:
                filters['patient_types'] = reg_df['patient_type'].dropna().unique().tolist()
        
        return filters

    def generate_excel_report(self, output_path: Optional[str] = None) -> bytes:
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self._write_filters_sheet(writer)
            self._write_overview_sheet(writer)
            self._write_anomalies_sheet(writer)
            self._write_department_sheet(writer)
            self._write_doctor_sheet(writer)
            self._write_wait_analysis_sheet(writer)
            self._write_fee_sheet(writer)
            self._write_satisfaction_sheet(writer)
            self._write_data_quality_sheet(writer)

        output.seek(0)
        self._format_excel(output)

        if output_path:
            with open(output_path, 'wb') as f:
                f.write(output.getvalue())

        return output.getvalue()

    def _format_excel(self, output: io.BytesIO):
        output.seek(0)
        wb = load_workbook(output)
        
        header_fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
        header_font = Font(bold=True, color='white')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)
        wb.save(output)

    def _write_filters_sheet(self, writer):
        filter_data = []
        
        if 'start_date' in self.filters and 'end_date' in self.filters:
            filter_data.append({
                '筛选维度': '日期范围',
                '筛选值': f"{self.filters['start_date']} 至 {self.filters['end_date']}"
            })
        
        if 'departments' in self.filters:
            dept_val = self.filters['departments']
            if isinstance(dept_val, list):
                dept_val = '、'.join(dept_val)
            filter_data.append({'筛选维度': '科室', '筛选值': dept_val})
        
        if 'doctors' in self.filters:
            doc_val = self.filters['doctors']
            if isinstance(doc_val, list):
                doc_val = '、'.join(doc_val)
            filter_data.append({'筛选维度': '医生', '筛选值': doc_val})
        
        if 'patient_types' in self.filters:
            pt_val = '、'.join(self.filters['patient_types'])
            filter_data.append({'筛选维度': '患者类型', '筛选值': pt_val})

        if filter_data:
            df = pd.DataFrame(filter_data)
            df.to_excel(writer, sheet_name='筛选条件', index=False)

    def _write_overview_sheet(self, writer):
        overview_metrics = self.metrics.get_overview_metrics()
        
        overview_data = []
        for key, value in overview_metrics.items():
            if isinstance(value, (int, float, str)):
                label = self._get_metric_label(key)
                overview_data.append({'核心指标': label, '数值': value})

        df = pd.DataFrame(overview_data)
        df.to_excel(writer, sheet_name='运营概览', index=False, startrow=0)
        
        filter_summary = []
        if 'start_date' in self.filters and 'end_date' in self.filters:
            filter_summary.append({'筛选条件': '日期范围', '值': f"{self.filters['start_date']} 至 {self.filters['end_date']}"})
        if 'departments' in self.filters:
            dept_val = self.filters['departments']
            if isinstance(dept_val, list):
                dept_val = f'共{len(dept_val)}个科室'
            filter_summary.append({'筛选条件': '科室', '值': dept_val})
        if 'patient_types' in self.filters:
            filter_summary.append({'筛选条件': '患者类型', '值': '、'.join(self.filters['patient_types'])})
        
        if filter_summary:
            filter_df = pd.DataFrame(filter_summary)
            filter_df.to_excel(writer, sheet_name='运营概览', index=False, startrow=len(df) + 3)

    def _write_anomalies_sheet(self, writer):
        anomaly_analysis = self.metrics.get_anomaly_cause_analysis()
        
        if anomaly_analysis:
            anomaly_data = []
            for item in anomaly_analysis:
                key_metrics = item.get('关键指标', {})
                metric_str = '; '.join([f"{self._get_metric_label(str(k))}: {v}" for k, v in key_metrics.items() 
                                       if k not in ['department_name', 'anomaly_type']])
                
                anomaly_data.append({
                    '科室名称': item.get('科室名称', ''),
                    '异常类型': item.get('异常类型', ''),
                    '可能原因': '; '.join(item.get('可能原因', [])),
                    '关键指标': metric_str
                })
            
            df = pd.DataFrame(anomaly_data)
            df.to_excel(writer, sheet_name='异常科室分析', index=False)

    def _write_department_sheet(self, writer):
        dept_metrics = self.metrics.get_department_metrics()
        if dept_metrics is not None:
            column_mapping = {
                'department_name': '科室名称',
                'total_registrations': '挂号量',
                'unique_patients': '就诊患者数',
                'total_visits': '就诊量',
                'doctor_count': '医生数',
                'visits_per_doctor': '人均接诊量',
                'avg_wait_time': '平均候诊时间(分钟)',
                'median_wait_time': '中位候诊时间(分钟)',
                'avg_satisfaction': '平均满意度'
            }
            df = dept_metrics.rename(columns=column_mapping)
            df = df[[col for col in column_mapping.values() if col in df.columns]]
            df.to_excel(writer, sheet_name='科室分析', index=False)

    def _write_doctor_sheet(self, writer):
        doc_metrics = self.metrics.get_doctor_metrics()
        if doc_metrics is not None:
            column_mapping = {
                'doctor_name': '医生姓名',
                'department_name': '所属科室',
                'title': '职称',
                'total_visits': '接诊量',
                'total_revenue': '收入贡献(元)',
                'avg_satisfaction': '平均满意度'
            }
            df = doc_metrics.rename(columns=column_mapping)
            df = df[[col for col in column_mapping.values() if col in df.columns]]
            df.to_excel(writer, sheet_name='医生分析', index=False)

    def _write_wait_analysis_sheet(self, writer):
        stratification = self.metrics.get_wait_time_stratification()
        if stratification:
            strata_data = []
            for strata, info in stratification.items():
                strata_data.append({
                    '候诊时间分层': strata,
                    '人数': info.get('人数', 0),
                    '占比(%)': info.get('占比', 0)
                })
            df_strata = pd.DataFrame(strata_data)
            df_strata.to_excel(writer, sheet_name='候诊分析', index=False, startrow=0)

        weekend_compare = self.metrics.get_weekday_weekend_comparison()
        if weekend_compare:
            compare_data = []
            for metric, values in weekend_compare.items():
                row = {'对比维度': metric}
                row.update(values)
                compare_data.append(row)
            df_compare = pd.DataFrame(compare_data)
            df_compare.to_excel(writer, sheet_name='候诊分析', index=False, startrow=len(df_strata) + 3 if stratification else 0)

        peak_hours = self.metrics.get_peak_hours()
        if peak_hours is not None:
            peak_data = []
            for _, row in peak_hours.iterrows():
                peak_data.append({
                    '时段(小时)': int(row['hour']),
                    '挂号量': row['registrations'],
                    '是否高峰时段': '是' if row['is_peak'] else '否'
                })
            df_peak = pd.DataFrame(peak_data)
            start_row = len(df_strata) + len(df_compare) + 6 if stratification and weekend_compare else len(df_strata) + 3 if stratification else len(df_compare) + 3 if weekend_compare else 0
            df_peak.to_excel(writer, sheet_name='候诊分析', index=False, startrow=start_row)

    def _write_fee_sheet(self, writer):
        fee_structure = self.metrics.get_fee_structure()

        summary_data = [
            {'项目': '总收入(元)', '金额': fee_structure.get('total_revenue', 0)},
            {'项目': '检查收入(元)', '金额': fee_structure.get('exam_revenue', 0)},
            {'项目': '药品收入(元)', '金额': fee_structure.get('drug_revenue', 0)},
            {'项目': '检查占比(%)', '金额': fee_structure.get('exam_ratio', 0)},
            {'项目': '药品占比(%)', '金额': fee_structure.get('drug_ratio', 0)}
        ]
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='费用分析', index=False, startrow=0)

        if 'exam_by_item' in fee_structure:
            exam_df = fee_structure['exam_by_item'].copy()
            exam_df.columns = ['检查项目', '总费用(元)', '检查次数']
            exam_df.to_excel(writer, sheet_name='费用分析', index=False, startrow=8)

        if 'drug_by_category' in fee_structure:
            drug_df = fee_structure['drug_by_category'].copy()
            drug_df.columns = ['药品类别', '总费用(元)']
            drug_start = 8 + len(exam_df) + 3 if 'exam_by_item' in fee_structure else 8
            drug_df.to_excel(writer, sheet_name='费用分析', index=False, startrow=drug_start)

        if self.transformed_data.get('visits') is not None and 'total_fee' in self.transformed_data['visits'].columns:
            visit_df = self.transformed_data['visits']
            fee_q3 = visit_df['total_fee'].quantile(0.75)
            fee_q1 = visit_df['total_fee'].quantile(0.25)
            fee_iqr = fee_q3 - fee_q1
            high_fee_threshold = fee_q3 + 1.5 * fee_iqr
            low_fee_threshold = max(0, fee_q1 - 1.5 * fee_iqr)
            
            high_fee_visits = visit_df[visit_df['total_fee'] > high_fee_threshold]
            low_fee_visits = visit_df[visit_df['total_fee'] < low_fee_threshold]
            
            anomaly_data = [
                {'异常类型': '高额费用记录', '数量': len(high_fee_visits), '阈值(元)': round(high_fee_threshold, 2)},
                {'异常类型': '低额费用记录', '数量': len(low_fee_visits), '阈值(元)': round(low_fee_threshold, 2)}
            ]
            anomaly_start = drug_start + len(drug_df) + 3 if 'drug_by_category' in fee_structure else 8 + len(exam_df) + 3 if 'exam_by_item' in fee_structure else 8
            pd.DataFrame(anomaly_data).to_excel(writer, sheet_name='费用分析', index=False, startrow=anomaly_start)

    def _write_satisfaction_sheet(self, writer):
        sat_dist = self.metrics.get_satisfaction_distribution()

        if 'avg_scores' in sat_dist:
            avg_data = []
            score_labels = {
                'overall_score': '总体满意度',
                'wait_score': '候诊满意度',
                'service_score': '服务满意度'
            }
            for k, v in sat_dist['avg_scores'].items():
                avg_data.append({'维度': score_labels.get(k, k), '平均分': v})
            pd.DataFrame(avg_data).to_excel(writer, sheet_name='满意度分析', index=False, startrow=0)

        strata_data = []
        for col in ['overall_score', 'wait_score', 'service_score']:
            dist_key = f'{col}_distribution'
            if dist_key in sat_dist:
                dist = sat_dist[dist_key]
                for score, count in dist.items():
                    strata_data.append({
                        '维度': score_labels.get(col, col),
                        '评分': int(score),
                        '人数': count
                    })
        if strata_data:
            pd.DataFrame(strata_data).to_excel(writer, sheet_name='满意度分析', index=False, startrow=len(avg_data) + 3)

        if 'recommendation_rate' in sat_dist:
            rec_data = [{'指标': '推荐率(%)', '数值': sat_dist['recommendation_rate']}]
            rec_start = len(avg_data) + len(strata_data) + 6 if strata_data else len(avg_data) + 3
            pd.DataFrame(rec_data).to_excel(writer, sheet_name='满意度分析', index=False, startrow=rec_start)

    def _write_data_quality_sheet(self, writer):
        validation_results = self.validator.validate_all()
        summary = self.validator.get_summary()
        
        summary_data = [
            {'项目': '错误总数', '数值': summary.get('total_errors', 0)},
            {'项目': '警告总数', '数值': summary.get('total_warnings', 0)},
            {'项目': '数据有效性', '数值': '通过' if summary.get('is_valid', False) else '未通过'}
        ]
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='数据质量', index=False, startrow=0)

        missing_data = []
        if 'missing_values' in validation_results:
            for table, info in validation_results['missing_values'].items():
                table_label = self.validator.FILE_LABELS.get(table, table)
                missing_data.append({
                    '数据表': table_label,
                    '缺失值总数': info.get('total_missing', 0),
                    '缺失率(%)': info.get('missing_percent', 0)
                })
        if missing_data:
            pd.DataFrame(missing_data).to_excel(writer, sheet_name='数据质量', index=False, startrow=len(summary_data) + 3)

        duplicate_data = []
        if 'duplicate_records' in validation_results:
            for table, issues in validation_results['duplicate_records'].items():
                table_label = self.validator.FILE_LABELS.get(table, table)
                duplicate_data.append({
                    '数据表': table_label,
                    '重复记录数': len(issues)
                })
        if duplicate_data:
            dup_start = len(summary_data) + len(missing_data) + 6 if missing_data else len(summary_data) + 3
            pd.DataFrame(duplicate_data).to_excel(writer, sheet_name='数据质量', index=False, startrow=dup_start)

        anomaly_data = []
        if 'anomaly_detection' in validation_results:
            for table, anomalies in validation_results['anomaly_detection'].items():
                table_label = self.validator.FILE_LABELS.get(table, table)
                for anomaly in anomalies:
                    anomaly_data.append({
                        '数据表': table_label,
                        '异常类型': anomaly.get('type', ''),
                        '描述': anomaly.get('description', '')
                    })
        if anomaly_data:
            anom_start = dup_start + len(duplicate_data) + 3 if duplicate_data else len(summary_data) + len(missing_data) + 6 if missing_data else len(summary_data) + 3
            pd.DataFrame(anomaly_data).to_excel(writer, sheet_name='数据质量', index=False, startrow=anom_start)

    def generate_html_report(self) -> str:
        overview = self.metrics.get_overview_metrics()
        dept_metrics = self.metrics.get_department_metrics()
        doc_metrics = self.metrics.get_doctor_metrics()
        fee_structure = self.metrics.get_fee_structure()
        sat_dist = self.metrics.get_satisfaction_distribution()
        anomaly_analysis = self.metrics.get_anomaly_cause_analysis()
        wait_stratification = self.metrics.get_wait_time_stratification()
        weekend_compare = self.metrics.get_weekday_weekend_comparison()
        peak_hours = self.metrics.get_peak_hours()
        validation_summary = self.validator.get_summary()
        missing_values = self.validator.validation_results.get('missing_values', {}) if hasattr(self.validator, 'validation_results') else {}
        mom_changes = self.metrics.get_mom_changes()

        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>医院门诊运营分析报告</title>
    <style>
        body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .header {{ background: linear-gradient(135deg, #1f77b4, #17a2b8); color: white; padding: 20px; border-radius: 10px; margin-bottom: 20px; }}
        .header h1 {{ margin: 0; font-size: 24px; }}
        .header p {{ margin: 5px 0 0 0; opacity: 0.9; }}
        .section {{ background: white; padding: 20px; border-radius: 10px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .section h2 {{ color: #1f77b4; border-bottom: 2px solid #1f77b4; padding-bottom: 10px; margin-top: 0; }}
        .section h3 {{ color: #2c3e50; font-size: 16px; margin-top: 20px; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }}
        .kpi-card {{ background: linear-gradient(135deg, #f8f9fa, #e9ecef); padding: 15px; border-radius: 8px; text-align: center; }}
        .kpi-value {{ font-size: 28px; font-weight: bold; color: #1f77b4; }}
        .kpi-label {{ font-size: 14px; color: #666; margin-top: 5px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background-color: #f8f9fa; font-weight: bold; color: #333; }}
        tr:hover {{ background-color: #f8f9fa; }}
        .footer {{ text-align: center; color: #666; margin-top: 30px; padding: 20px; }}
        .highlight {{ background-color: #fff3cd; }}
        .warning {{ background-color: #f8d7da; }}
        .success {{ background-color: #d4edda; }}
        .filter-badge {{ display: inline-block; background: #e3f2fd; color: #1565c0; padding: 5px 12px; border-radius: 20px; margin: 3px; font-size: 13px; }}
        .mom-positive {{ color: #28a745; }}
        .mom-negative {{ color: #dc3545; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🏥 医院门诊运营分析报告</h1>
        <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>

    <div class="section">
        <h2>🔍 筛选条件</h2>
        <div>
            {self._generate_filter_badges()}
        </div>
    </div>

    <div class="section">
        <h2>📊 运营概览</h2>
        <div class="kpi-grid">
            <div class="kpi-card">
                <div class="kpi-value">{overview.get('total_registrations', 0)}</div>
                <div class="kpi-label">总挂号量</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">{overview.get('unique_patients', 0)}</div>
                <div class="kpi-label">就诊患者数</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">¥{overview.get('total_revenue', 0):,.0f}</div>
                <div class="kpi-label">总收入</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">{overview.get('avg_wait_time', 0)}分钟</div>
                <div class="kpi-label">平均候诊时间</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">{overview.get('avg_overall_score', 0)}</div>
                <div class="kpi-label">平均满意度</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">{overview.get('total_departments', 0)}</div>
                <div class="kpi-label">科室数量</div>
            </div>
        </div>
    </div>

    <div class="section">
        <h2>📈 同比环比分析</h2>
        {self._generate_mom_html(mom_changes)}
    </div>

    <div class="section">
        <h2>⚠️ 异常科室分析</h2>
        {self._generate_anomaly_html(anomaly_analysis)}
    </div>

    <div class="section">
        <h2>⏱️ 候诊时间分析</h2>
        <h3>候诊时间分层</h3>
        {self._generate_wait_strata_html(wait_stratification)}
        <h3>工作日周末对比</h3>
        {self._generate_weekend_html(weekend_compare)}
        <h3>高峰时段</h3>
        {self._generate_peak_hours_html(peak_hours)}
    </div>

    <div class="section">
        <h2>🏥 科室运营分析</h2>
        <table>
            <thead>
                <tr>
                    <th>科室名称</th>
                    <th>挂号量</th>
                    <th>就诊量</th>
                    <th>医生数</th>
                    <th>人均接诊量</th>
                    <th>平均候诊时间</th>
                    <th>平均满意度</th>
                </tr>
            </thead>
            <tbody>
                {self._generate_dept_rows(dept_metrics)}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>👨‍⚕️ 医生绩效Top 10</h2>
        <table>
            <thead>
                <tr>
                    <th>排名</th>
                    <th>医生姓名</th>
                    <th>科室</th>
                    <th>职称</th>
                    <th>接诊量</th>
                    <th>收入贡献</th>
                    <th>满意度</th>
                </tr>
            </thead>
            <tbody>
                {self._generate_doctor_rows(doc_metrics)}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>💰 费用结构分析</h2>
        <div class="kpi-grid">
            <div class="kpi-card">
                <div class="kpi-value">¥{fee_structure.get('exam_revenue', 0):,.0f}</div>
                <div class="kpi-label">检查收入 ({fee_structure.get('exam_ratio', 0)}%)</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">¥{fee_structure.get('drug_revenue', 0):,.0f}</div>
                <div class="kpi-label">药品收入 ({fee_structure.get('drug_ratio', 0)}%)</div>
            </div>
        </div>
    </div>

    <div class="section">
        <h2>😊 满意度分析</h2>
        {self._generate_satisfaction_html(sat_dist)}
    </div>

    <div class="section">
        <h2>✅ 数据质量摘要</h2>
        {self._generate_data_quality_html(validation_summary, missing_values)}
    </div>

    <div class="footer">
        <p>本报告由医院门诊运营分析平台自动生成</p>
    </div>
</body>
</html>
        """
        return html_content

    def _generate_filter_badges(self) -> str:
        badges = []
        if 'start_date' in self.filters and 'end_date' in self.filters:
            badges.append(f'<span class="filter-badge">📅 {self.filters["start_date"]} 至 {self.filters["end_date"]}</span>')
        if 'departments' in self.filters:
            dept_val = self.filters['departments']
            if isinstance(dept_val, list):
                dept_val = f'共{len(dept_val)}个科室'
            badges.append(f'<span class="filter-badge">🏥 {dept_val}</span>')
        if 'patient_types' in self.filters:
            badges.append(f'<span class="filter-badge">👤 {"、".join(self.filters["patient_types"])}</span>')
        return ''.join(badges) if badges else '<span style="color: #999;">无筛选条件</span>'

    def _generate_mom_html(self, mom_changes: Dict[str, Any]) -> str:
        if not mom_changes:
            return '<p>暂无环比数据</p>'
        
        html = '<div class="kpi-grid">'
        for metric, data in mom_changes.items():
            if data and len(data) >= 2:
                latest = data[-1]
                mom_rate = latest.get('环比变化率')
                if mom_rate is not None:
                    rate_class = 'mom-positive' if mom_rate >= 0 else 'mom-negative'
                    rate_sign = '+' if mom_rate >= 0 else ''
                    display_value = latest.get('数量')
                    if display_value is None:
                        display_value = latest.get('金额')
                    if display_value is None:
                        display_value = latest.get('平均候诊时间')
                    if display_value is None:
                        display_value = latest.get('平均满意度', '')
                    html += f'''
                    <div class="kpi-card">
                        <div class="kpi-value">{display_value}</div>
                        <div class="kpi-label">{metric} <span class="{rate_class}">({rate_sign}{mom_rate}%)</span></div>
                    </div>
                    '''
        html += '</div>'
        return html

    def _generate_anomaly_html(self, anomaly_analysis: List[Dict]) -> str:
        if not anomaly_analysis:
            return '<p>暂无异常科室</p>'
        
        rows = []
        for item in anomaly_analysis:
            rows.append(f'''
                <tr class="warning">
                    <td>{item.get("科室名称", "")}</td>
                    <td>{item.get("异常类型", "")}</td>
                    <td>{"; ".join(item.get("可能原因", []))}</td>
                </tr>
            ''')
        
        return f'''
            <table>
                <thead>
                    <tr>
                        <th>科室名称</th>
                        <th>异常类型</th>
                        <th>可能原因</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
        '''

    def _generate_wait_strata_html(self, stratification: Dict[str, Any]) -> str:
        if not stratification:
            return '<p>暂无候诊时间数据</p>'
        
        rows = []
        for strata, info in stratification.items():
            rows.append(f'''
                <tr>
                    <td>{strata}</td>
                    <td>{info.get("人数", 0)}</td>
                    <td>{info.get("占比", 0)}%</td>
                </tr>
            ''')
        
        return f'''
            <table>
                <thead>
                    <tr>
                        <th>候诊时间分层</th>
                        <th>人数</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
        '''

    def _generate_weekend_html(self, weekend_compare: Dict[str, Any]) -> str:
        if not weekend_compare:
            return '<p>暂无工作日周末对比数据</p>'
        
        rows = []
        for metric, values in weekend_compare.items():
            row = f'<td>{metric}</td>'
            for k, v in values.items():
                row += f'<td>{v}</td>'
            rows.append(f'<tr>{row}</tr>')
        
        headers = ['对比维度'] + list(next(iter(weekend_compare.values())).keys())
        return f'''
            <table>
                <thead>
                    <tr>
                        {"".join([f"<th>{h}</th>" for h in headers])}
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
        '''

    def _generate_peak_hours_html(self, peak_hours: Optional[pd.DataFrame]) -> str:
        if peak_hours is None or len(peak_hours) == 0:
            return '<p>暂无高峰时段数据</p>'
        
        rows = []
        for _, row in peak_hours.iterrows():
            is_peak = row.get('is_peak', False)
            row_class = 'highlight' if is_peak else ''
            rows.append(f'''
                <tr class="{row_class}">
                    <td>{int(row["hour"])}:00 - {int(row["hour"]) + 1}:00</td>
                    <td>{row["registrations"]}</td>
                    <td>{'⭐ 高峰' if is_peak else '常规'}</td>
                </tr>
            ''')
        
        return f'''
            <table>
                <thead>
                    <tr>
                        <th>时段</th>
                        <th>挂号量</th>
                        <th>类型</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
        '''

    def _generate_satisfaction_html(self, sat_dist: Dict[str, Any]) -> str:
        if not sat_dist or 'avg_scores' not in sat_dist:
            return '<p>暂无满意度数据</p>'
        
        score_labels = {
            'overall_score': '总体满意度',
            'wait_score': '候诊满意度',
            'service_score': '服务满意度'
        }
        
        avg_rows = []
        for k, v in sat_dist['avg_scores'].items():
            avg_rows.append(f'''
                <tr>
                    <td>{score_labels.get(k, k)}</td>
                    <td>{v}</td>
                </tr>
            ''')
        
        html = f'''
            <h3>各维度平均分</h3>
            <table>
                <thead>
                    <tr>
                        <th>维度</th>
                        <th>平均分</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(avg_rows)}
                </tbody>
            </table>
        '''
        
        if 'recommendation_rate' in sat_dist:
            html += f'''
                <div class="kpi-card" style="margin-top: 20px; display: inline-block;">
                    <div class="kpi-value">{sat_dist["recommendation_rate"]}%</div>
                    <div class="kpi-label">患者推荐率</div>
                </div>
            '''
        
        return html

    def _generate_data_quality_html(self, summary: Dict[str, Any], missing_values: Dict[str, Any]) -> str:
        is_valid = summary.get('is_valid', False)
        valid_class = 'success' if is_valid else 'warning'
        
        html = f'''
            <div class="kpi-grid">
                <div class="kpi-card">
                    <div class="kpi-value {valid_class}">{summary.get('total_errors', 0)}</div>
                    <div class="kpi-label">错误数</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-value">{summary.get('total_warnings', 0)}</div>
                    <div class="kpi-label">警告数</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-value {'success' if is_valid else 'warning'}">{'通过' if is_valid else '未通过'}</div>
                    <div class="kpi-label">数据有效性</div>
                </div>
            </div>
        '''
        
        if missing_values:
            rows = []
            for table, info in missing_values.items():
                table_label = self.validator.FILE_LABELS.get(table, table)
                rows.append(f'''
                    <tr>
                        <td>{table_label}</td>
                        <td>{info.get('total_missing', 0)}</td>
                        <td>{info.get('missing_percent', 0)}%</td>
                    </tr>
                ''')
            
            html += f'''
                <h3>缺失值统计</h3>
                <table>
                    <thead>
                        <tr>
                            <th>数据表</th>
                            <th>缺失值总数</th>
                            <th>缺失率</th>
                        </tr>
                    </thead>
                    <tbody>
                        {''.join(rows)}
                    </tbody>
                </table>
            '''
        
        return html

    def _generate_dept_rows(self, dept_metrics):
        if dept_metrics is None or len(dept_metrics) == 0:
            return '<tr><td colspan="7" style="text-align:center;">暂无数据</td></tr>'

        rows = []
        for _, row in dept_metrics.iterrows():
            rows.append(f"""
                <tr>
                    <td>{row.get('department_name', '')}</td>
                    <td>{row.get('total_registrations', 0)}</td>
                    <td>{row.get('total_visits', 0)}</td>
                    <td>{row.get('doctor_count', 0)}</td>
                    <td>{row.get('visits_per_doctor', 0)}</td>
                    <td>{row.get('avg_wait_time', 0):.1f}分钟</td>
                    <td>{row.get('avg_satisfaction', 0):.2f}</td>
                </tr>
            """)
        return ''.join(rows)

    def _generate_doctor_rows(self, doc_metrics):
        if doc_metrics is None or len(doc_metrics) == 0:
            return '<tr><td colspan="7" style="text-align:center;">暂无数据</td></tr>'

        rows = []
        for idx, (_, row) in enumerate(doc_metrics.head(10).iterrows(), 1):
            rows.append(f"""
                <tr>
                    <td>{idx}</td>
                    <td>{row.get('doctor_name', '')}</td>
                    <td>{row.get('department_name', '')}</td>
                    <td>{row.get('title', '')}</td>
                    <td>{row.get('total_visits', 0)}</td>
                    <td>¥{row.get('total_revenue', 0):,.0f}</td>
                    <td>{row.get('avg_satisfaction', 0):.2f}</td>
                </tr>
            """)
        return ''.join(rows)

    def _get_metric_label(self, key: str) -> str:
        labels = {
            'total_registrations': '总挂号量',
            'unique_patients': '就诊患者数',
            'avg_daily_registrations': '日均挂号量',
            'total_visits': '总就诊量',
            'total_revenue': '总收入',
            'avg_visit_fee': '次均费用',
            'avg_wait_time': '平均候诊时间(分钟)',
            'median_wait_time': '中位候诊时间(分钟)',
            'max_wait_time': '最长候诊时间(分钟)',
            'wait_over_30min_pct': '候诊超30分钟占比(%)',
            'avg_overall_score': '平均满意度',
            'satisfaction_pct': '满意度达标率(%)',
            'total_departments': '科室数量',
            'total_doctors': '医生数量',
            'doctor_count': '医生数',
            'visits_per_doctor': '人均接诊量',
            'avg_satisfaction': '平均满意度',
            'anomaly_type': '异常类型',
            'department_name': '科室名称'
        }
        return labels.get(key, key)

    def save_html_report(self, output_path: str) -> None:
        html_content = self.generate_html_report()
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def get_report_filename(self, report_type: str = 'excel') -> str:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if report_type == 'excel':
            return f'门诊运营分析报告_{timestamp}.xlsx'
        elif report_type == 'html':
            return f'门诊运营分析报告_{timestamp}.html'
        return f'门诊运营分析报告_{timestamp}'
